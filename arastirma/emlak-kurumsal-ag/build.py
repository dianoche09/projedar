#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Emlak Kurumsal Ağ araştırma veri setini CSV + XLSX olarak derler.

Kullanım:  /tmp/.venv-xlsx/bin/python build.py
"""
import csv
import json
import os

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")
OUT = os.path.join(BASE, "cikti")
TARIH = "2026-08-13"

os.makedirs(OUT, exist_ok=True)


def load(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


ANA_BASLIK = [
    "Kuruluş Adı", "Kısaltma", "Kuruluş Türü", "Ulusal/Bölgesel/Yerel", "İl", "İlçe",
    "Adres", "Telefon", "WhatsApp", "E-posta", "Web Sitesi",
    "Instagram", "LinkedIn", "Facebook", "X/Twitter", "YouTube",
    "Başkan", "Genel Sekreter/Yönetici", "Üye Sayısı", "Üye Sayısı Kaynak Tarihi",
    "Üye Profili", "Faaliyet Alanı", "Bağlı Olduğu Üst Kuruluş", "Alt Kuruluş/Şube",
    "Etkinlik Düzenliyor mu?", "Eğitim Veriyor mu?", "Üyelik Başvuru Linki",
    "İşbirliği/Protokol Bilgisi", "Dijital Aktiflik", "İşbirliği Potansiyel Skoru",
    "Skor Gerekçesi", "Önerilen İşbirliği Modeli", "Aktiflik Durumu",
    "Kaynak Linkleri", "Son Doğrulama Tarihi", "Notlar",
]

ALAN = ["ad", "kisaltma", "tur", "olcek", "il", "ilce", "adres", "telefon", "whatsapp",
        "eposta", "web", "instagram", "linkedin", "facebook", "x", "youtube",
        "baskan", "yonetici", "uye_sayisi", "uye_kaynak_tarihi", "uye_profili",
        "faaliyet", "ust_kurulus", "alt_kurulus", "etkinlik", "egitim", "uyelik_link",
        "protokol", "dijital", "skor", "skor_gerekce", "model", "aktiflik", "kaynak"]

# --- Ana kuruluş listesi -----------------------------------------------------
kuruluslar = []
for f in ("01-ulusal.json", "02-sektorel-stk.json", "03a-il-odalari.json",
          "03b-il-odalari.json", "04a-dernekler-detayli.json"):
    kuruluslar.extend(load(f))

# TEMFED üye dernekleri (ad/il/ilçe/skor/not formatında kısa kayıtlar)
D = "Doğrulanamadı"
for ad, il, ilce, skor, not_ in load("04b-dernekler-temfed-liste.json"):
    kuruluslar.append({
        "ad": ad, "kisaltma": "-", "tur": "Dernek",
        "olcek": "Yerel (ilçe)" if ilce not in ("Merkez", "-") else "Yerel (il)",
        "il": il, "ilce": ilce, "adres": D, "telefon": D, "whatsapp": D, "eposta": D,
        "web": D, "instagram": D, "linkedin": D, "facebook": D, "x": D, "youtube": D,
        "baskan": D, "yonetici": "-", "uye_sayisi": "Açıklanmamış", "uye_kaynak_tarihi": "-",
        "uye_profili": "Yerel emlak müşavirleri / komisyoncuları",
        "faaliyet": "Mesleki temsil (detay doğrulanmadı)",
        "ust_kurulus": "TEMFED üyesi", "alt_kurulus": "-",
        "etkinlik": D, "egitim": D, "uyelik_link": "-", "protokol": D, "dijital": D,
        "skor": skor,
        "skor_gerekce": "TEMFED üye listesinde doğrulanmış tüzel kişilik; başkan/iletişim/üye "
                        "verisi bu araştırmada elde edilemedi. Skor pazar hacmi tahminine dayanır, "
                        "kurumsal kapasiteye değil.",
        "model": "Üyelere ücretsiz deneme (önce kurumsal doğrulama)",
        "aktiflik": "TEMFED üye listesinde yer alıyor; bağımsız faaliyet kanıtı doğrulanmadı",
        "kaynak": "https://www.temfed.org.tr/meslek-oda-ve-dernek-uyeleri",
        "notlar": not_,
    })

kuruluslar.sort(key=lambda r: (-int(r["skor"]), r["ad"]))


def yaz_csv(dosya, baslik, satirlar):
    p = os.path.join(OUT, dosya)
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(baslik)
        w.writerows(satirlar)
    return p


ana_satirlar = [[k.get(a, D) for a in ALAN] + [TARIH, k.get("notlar", "")]
                for k in kuruluslar]
yaz_csv("01-emlak-kuruluslari.csv", ANA_BASLIK, ana_satirlar)

ek = load("05-ek-sheetler.json")

KOM_B = ["Oda", "Kısaltma", "Grup/Komite No", "Komite Adı", "İl", "Komite Başkanı",
         "Başkan Vekili", "Komite Üyeleri", "Grup Üye Sayısı", "Web",
         "İşbirliği Skoru", "Skor Gerekçesi", "Önerilen Model", "Kaynak", "Not"]
kom = [[c["oda"], c["kisaltma"], c["grup_no"], c["komite"], c["il"], c["baskan"],
        c["baskan_yrd"], c["uyeler"], c["grup_uye_sayisi"], c["web"], c["skor"],
        c["gerekce"], c["model"], c["kaynak"], c["not"]]
       for c in ek["ticaret_odasi_komiteleri"]]
yaz_csv("02-ticaret-odasi-komiteleri.csv", KOM_B, kom)

KIS_B = ["Kişi", "Roller", "Neden Öncelikli", "Son Doğrulama"]
kis = [[k["kisi"], k["roller"], k["neden_onemli"], TARIH] for k in ek["kesisim_kisiler"]]
yaz_csv("03-yonetim-karar-vericiler.csv", KIS_B, kis)

PRO_B = ["Kuruluş", "İşbirliği Yapılan Kurum", "İşbirliği Türü", "Tarih", "Açıklama", "Kaynak"]
pro = [[p["kurulus"], p["partner"], p["tur"], p["tarih"], p["aciklama"], p["kaynak"]]
       for p in ek["protokol_gecmisi"]]
yaz_csv("04-protokol-gecmisi.csv", PRO_B, pro)

FRA_B = ["Marka", "Tür", "Merkez", "Kuruluş/Devir", "Ofis Sayısı", "Danışman Sayısı",
         "Kapsam", "İşbirliği Skoru", "Skor Gerekçesi", "Önerilen Model", "Kaynak", "Not"]
fra = [[f["marka"], f["tur"], f["merkez"], f["kurulus"], f["ofis"], f["danisman"],
        f["kapsam"], f["skor"], f["gerekce"], f["model"], f["kaynak"], f["not"]]
       for f in ek["franchise_yapilar"]]
yaz_csv("05-franchise-yapilar.csv", FRA_B, fra)

AG_B = ["Seviye 1 (Üst Kuruluş)", "Seviye 2", "Seviye 3", "Seviye 4"]
yaz_csv("06-kurumsal-ag.csv", AG_B, ek["kurumsal_ag"])

# --- Top liste (skora göre, tüm kayıtlar dahil) ------------------------------
TOP_B = ["Sıra", "Kuruluş", "Tür", "İl/Bölge", "Üye Sayısı", "Başkan/Yetkili",
         "Telefon", "E-posta", "Web Sitesi", "Skor", "Önerilen İşbirliği Modeli",
         "Neden Öncelikli?"]
top = []
sira = 0
for k in kuruluslar:
    sira += 1
    top.append([sira, k["ad"], k["tur"], f"{k['il']} / {k['olcek']}", k["uye_sayisi"],
                k["baskan"], k["telefon"], k["eposta"], k["web"], k["skor"],
                k["model"], k["skor_gerekce"]])
# franchise + komiteler de öncelik listesine
for c in ek["ticaret_odasi_komiteleri"]:
    sira += 1
    top.append([sira, f"{c['oda']} - {c['komite']}", "Ticaret odası meslek komitesi",
                c["il"], c["grup_uye_sayisi"], c["baskan"], "-", "-", c["web"],
                c["skor"], c["model"], c["gerekce"]])
for f in ek["franchise_yapilar"]:
    sira += 1
    top.append([sira, f["marka"], f["tur"], f["merkez"],
                f"{f['ofis']} ofis / {f['danisman']} danışman", "-", "-", "-",
                f["kaynak"].split(" | ")[0], f["skor"], f["model"], f["gerekce"]])
top.sort(key=lambda r: -int(r[9]))
for i, r in enumerate(top, 1):
    r[0] = i
yaz_csv("07-kurumsal-oncelik-listesi.csv", TOP_B, top)

# --- XLSX --------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("Emlak Kuruluşları", ANA_BASLIK, ana_satirlar),
        ("Ticaret Odası Komiteleri", KOM_B, kom),
        ("Yönetim-Karar Vericiler", KIS_B, kis),
        ("Protokol Geçmişi", PRO_B, pro),
        ("Franchise Yapılar", FRA_B, fra),
        ("Kurumsal Ağ", AG_B, ek["kurumsal_ag"]),
        ("Kurumsal Öncelik Listesi", TOP_B, top),
    ]
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    for ad, baslik, satirlar in sheets:
        ws = wb.create_sheet(ad[:31])
        ws.append(baslik)
        for s in satirlar:
            ws.append(s)
        for c in ws[1]:
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i in range(1, len(baslik) + 1):
            uzunluklar = [len(str(ws.cell(r, i).value or "")) for r in range(1, min(ws.max_row, 60) + 1)]
            ws.column_dimensions[get_column_letter(i)].width = min(max(max(uzunluklar) + 2, 12), 55)
        ws.row_dimensions[1].height = 32
    wb.save(os.path.join(OUT, "Turkiye-Emlak-Kurumsal-Ag.xlsx"))
    print("XLSX yazıldı")
except ImportError:
    print("openpyxl yok - sadece CSV üretildi")

print(f"Toplam kuruluş kaydı : {len(kuruluslar)}")
print(f"Öncelik listesi satır: {len(top)}")
print(f"Protokol kaydı       : {len(pro)}")
print(f"Çıktı dizini         : {OUT}")
